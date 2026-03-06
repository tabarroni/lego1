import openpyxl
import os
from tkinter import messagebox

EXCEL_FILE = "inventario_lego.xlsx"

# 📘 Crear Excel si no existe
if not os.path.exists(EXCEL_FILE):
    wb = openpyxl.Workbook()
    hoja = wb.active
    hoja.title = "Piezas"
    hoja.append(["ID Pieza", "Descripción", "Color", "Cantidad", "Set Origen", "Ubicación"])
    wb.save(EXCEL_FILE)

def agregar_a_excel(id_pieza, descripcion, color, cantidad, set_origen, ubicacion):
    if not id_pieza or not descripcion or not str(cantidad).isdigit():
        messagebox.showerror("Error", "ID, Descripción y Cantidad válida son obligatorios.")
        return

    cantidad = int(cantidad)
    wb = openpyxl.load_workbook(EXCEL_FILE)

    # ✅ Asegurar que exista la hoja "Piezas"
    if "Piezas" not in wb.sheetnames:
        hoja = wb.create_sheet("Piezas")
        hoja.append(["ID Pieza", "Descripción", "Color", "Cantidad", "Set Origen", "Ubicación"])
    else:
        hoja = wb["Piezas"]

    # 🔍 Buscar si la pieza ya existe (por ID)
    pieza_existente = None
    for fila in hoja.iter_rows(min_row=2, values_only=False):
        id_celda = fila[0].value
        if str(id_celda) == str(id_pieza):
            pieza_existente = fila
            break

    if pieza_existente:
        # ✅ Si existe, actualiza la cantidad sumando
        celda_cantidad = pieza_existente[3]  # columna D (Cantidad)
        cantidad_actual = celda_cantidad.value if celda_cantidad.value else 0
        nueva_cantidad = cantidad_actual + cantidad
        celda_cantidad.value = nueva_cantidad
        messagebox.showinfo(
            "Actualizado",
            f"La pieza '{descripcion}' ya existía. Cantidad actualizada a {nueva_cantidad}."
        )
    else:
        # 🆕 Si no existe, agregar nueva fila
        hoja.append([id_pieza, descripcion, color, cantidad, set_origen, ubicacion])
        messagebox.showinfo("Éxito", f"Pieza '{descripcion}' agregada correctamente.")

    wb.save(EXCEL_FILE)

